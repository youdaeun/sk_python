#앞에서 진행했던 malwares 트래픽 사이트의 결과를 엑셀 파일로 저장하시오.
#1. 일자, 제목, 링크를 헤더로 입력
#2. 데이터를 저장

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


# URL 설정
url = "https://www.malware-traffic-analysis.net/2023/index.html"
headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}

# 요청 보내기
req = requests.get(url, headers=headers)
# BeautifulSoup로 HTML 파싱
soup = BeautifulSoup(req.text, "lxml")

wb = Workbook()
ws = wb.active
ws['A1'] ="No"
ws['B1']="설명"
ws['C1'] = "URL 링크"

# 제목과 링크 추출
tags = soup.select("#main_content > div.content > ul > li > a.main_menu")

i=2
# 결과 출력
for tag in tags:
    ws.cell(row=i,column=1, value=i-1) 
    ws.cell(row=i, column=2, value=tag.text)
    ws.cell(row=i, column=3, value=f"https://www.malware-traffic-analysis.net/2024/{tag['href']}")
    i = i+1

wb.save("malwares.xlsx")

