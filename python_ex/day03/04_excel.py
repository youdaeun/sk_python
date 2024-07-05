from openpyxl import Workbook, load_workbook

wb = Workbook()

save_path = "02_엑셀 자동화.xlsx"
# 기존 엑셀 파일 불러오기
ws = wb.create_sheet("SK 쉴더스 루키즈")

# 활성화된 시트 선택
ws = wb.active

# 데이터 추가(1)
# 셀 이름 지정해서 셀에 직접 데이터 추가
ws['A1'] = '날짜'
ws['B1'] = '제품명'
ws['C1'] = '가격'
ws['D1'] = '수량'
ws['E1'] = '합계'

# 데이터 추가(2)
ws.cell(row=2, column=1, value='2024-05-07')
ws.cell(row=2, column=2, value='아이패드')
ws.cell(row=2, column=3, value=780000)
ws.cell(row=2, column=4, value=30)
# 엑셀 수식
ws.cell(row=2, column=5, value='=C2*D2')

# 엑셀 저장
wb.save(save_path)