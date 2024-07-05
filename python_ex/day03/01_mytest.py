import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

#URL 설정
url = "https://www.yogiyo.co.kr/mobile/#/%EC%84%9C%EC%9A%B8%ED%8A%B9%EB%B3%84%EC%8B%9C/000000/"
headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}

req = requests.get(url, headers=headers)
#BeautifulSoup로 HTML 파싱
soup = BeautifulSoup(req.text,"lxml")

wb = Workbook()
ws = wb.active
ws['A1']="No."
ws['B1']="제목"
ws['C1']="이미지 링크"

#제목과 이미지 링크 추출
tags= soup.append("#content > div > div > div > div.restaurant-list > div > div > table > tbody > tr > td > div ")
i=2
for tag in tags:
    ws.cell(row=i, column=1, value=i-1)
    ws.cell(row=i, column=2, value=tag.text)
    i = i+1

wb.save("01_mytest.xlsx")