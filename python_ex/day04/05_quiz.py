import os
import time
import re
from datetime import datetime
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from openpyxl import Workbook

# 정규 표현식 패턴 설정
email_pattern = re.compile(r'[\w\.-]+@[\w\.-]+')  # 이메일 패턴
jumin_pattern = re.compile(r'\d{6}[-]\d{7}')     # 주민등록번호 패턴

# Slack API 토큰 및 채널 설정
SLACK_API_TOKEN = "your_slack_api_token"
SLACK_CHANNEL ="your_slack_channel"

DIR_PATH = 'static'  # 감시할 디렉토리 경로

# Slack WebClient 초기화
client = WebClient(token=SLACK_API_TOKEN)

def send_message(channel, text):
    try:
        response = client.chat_postMessage(
            channel=channel,
            text=text
        )
        print("메시지 전송 성공:", response["message"]["text"])
    except SlackApiError as e:
        print("메시지 전송 실패:", e.response["error"])

def process_php_file(file_path):
    # Feed로 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "PHP File Analysis"

    row = 1  # 엑셀의 시작 행 번호

    # 파일 내용 읽기
    with open(file_path, "r", encoding="UTF-8") as file:
        lines = file.readlines()
        for num, line in enumerate(lines, start=1):
            # 이메일 탐지
            emails = email_pattern.findall(line)
            for email in emails:
                send_message(SLACK_CHANNEL, f"이메일 발견 - {num}번째 줄: {email}")
                ws[f'A{row}'] = email
                ws[f'B{row}'] = ''  # 주민등록번호 발견이 아니므로 빈 값으로 처리
                row += 1

            # 주민등록번호 탐지
            jumins = jumin_pattern.findall(line)
            for jumin in jumins:
                send_message(SLACK_CHANNEL, f"주민등록번호 발견 - {num}번째 줄: {jumin}")
                ws[f'B{row}'] = jumin
                ws[f'A{row}'] = ''  # 이메일 발견이 아니므로 빈 값으로 처리
                row += 1

    # 파일 저장
    excel_file_name = f"{os.path.splitext(os.path.basename(file_path))[0]}.xlsx"
    excel_file_path = os.path.join(DIR_PATH, excel_file_name)
    wb.save(excel_file_path)

    # 파일 업로드 경고 메시지 전송
    message = f"경고: 새로운 .php 파일이 디렉토리에 업로드되었습니다!\n파일: {excel_file_name}\n시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    send_message(SLACK_CHANNEL, message)

# 루프 시작 전 초기 파일 목록 설정
pre_file = set(os.listdir(DIR_PATH))

while True:
    # 현재 디렉토리의 파일 목록 설정
    current_file = set(os.listdir(DIR_PATH))
    # 이전 파일 목록과 현재 파일 목록의 차이 계산
    result_diff = current_file - pre_file

    for file_name in result_diff:
        if file_name.endswith(".php"):
            print(f"새로운 PHP 파일 감지: {file_name}")
            file_path = os.path.join(DIR_PATH, file_name)
            process_php_file(file_path)

    # 다음 반복을 위해 이전 파일 목록 업데이트
    pre_file = current_file
    print("새로운 파일 확인 중...")
    time.sleep(1)
