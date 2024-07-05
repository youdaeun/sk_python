#20기 스케줄링 코드
import requests
from bs4 import BeautifulSoup
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication 
from dotenv import load_dotenv
import os
import openpyxl
from datetime import datetime
import schedule
import time

def mail_sender(file_path, now):
        load_dotenv()
        SECRET_ID = os.getenv("ID")
        SECRET_PASS = os.getenv("PASS")

        smtp = smtplib.SMTP('smtp.naver.com', 587)
        smtp.ehlo()
        smtp.starttls()

        smtp.login(SECRET_ID,SECRET_PASS)

        myemail = "youremail"
        youremail = "youremail"

        msg = MIMEMultipart()

        msg['Subject'] =f"보안 동향 {now} 정보입니다."
        msg['From'] = myemail
        msg['To'] = youremail

        text = """
        <html>
        <body>
        <h2>보안 동향 {} 정보입니다.</h2>
        </body>
        </html>
        """.format(now)

        contentPart = MIMEText(text, "html") 
        msg.attach(contentPart) 

        etc_file_path = file_path
        with open(etc_file_path, 'rb') as f : 
                etc_part = MIMEApplication( f.read() )
                etc_part.add_header('Content-Disposition','attachment', filename=etc_file_path)
                msg.attach(etc_part)

        smtp.sendmail( myemail,youremail,msg.as_string() )
        smtp.quit()

def job():
        now = datetime.now().strftime("%Y-%m-%d")
        url = "https://www.malware-traffic-analysis.net/2023/index.html"

        headers = {
                'User-Agent': 'Mozilla/5.0',
                'Content-Type': 'text/html; charset=utf-8'
        }

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = "설명"
        worksheet['B1'] = "URL링크"

        req = requests.get(url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")

        tags = soup.select("#main_content > div.content > ul > li > a.main_menu")

        row = 2
        for tag in tags:
                tag_text = tag.text
                tag_href = f"https://www.malware-traffic-analysis.net/2023/{tag['href']}"
                worksheet.cell(row=row, column=1, value=tag_text)
                worksheet.cell(row=row, column=2, value=tag_href)
                row = row + 1

        workbook.save(f'malware_{now}.xlsx')
        #메일보내기
        mail_sender(f'malware_{now}.xlsx', now)

#스케줄링
schedule.every(1).days.do(job)

while True:
    schedule.run_pending()
    time.sleep(1)